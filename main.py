import calendar
import sqlite3
import datetime
import openpyxl
from datetime import datetime as www
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

from sortMasiv import sortMasiv

from PyQt5 import QtWidgets, uic, QtCore
from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import QMessageBox, QHBoxLayout, QWidget, QApplication, QTableWidgetItem

Form, _ = uic.loadUiType("ui_main.ui")


class MyWin(QtWidgets.QMainWindow, Form):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.setupUi(self)
        self.openProg = False
        self.conn = sqlite3.connect("BDGrafik.db")
        self.cursor = self.conn.cursor()
        self.TypeProducts = []
        self.Volume = []
        self.AutoclavaNumber = []
        self.AutoclavaFoo = []
        self.IssuanceManufacture = []
        self.IssuanceManufactureQ = []
        self.IssuanceManufactureA = []
        self.IssuanceTesting = []
        self.IssuanceTestingQ = []
        self.IssuanceTestingA = []
        self.IssuanceProtocol = []
        self.IssuanceProtocolQ = []
        self.IssuanceProtocolA = []
        self.DateSettingCycles = []
        self.DateDeliverySGP = []
        self.Note = []
        self.GT = []
        if int(datetime.date.today().month) != 12:
            self.comboBox.setCurrentIndex(int(datetime.date.today().month))
        else:
            self.comboBox.setCurrentIndex(0)
        year = str(datetime.date.today().year)
        dateEditYear = QtCore.QDate.fromString(year, "yyyy")
        self.dateEdit.setDate(dateEditYear)
        self.pushButton_2.clicked.connect(self.format)
        self.pushButton_3.clicked.connect(self.addZapis)
        self.pushButton_4.clicked.connect(self.delZapis)
        self.pushButton.clicked.connect(self.save)
        self.comboBox.activated.connect(self.open)
        self.dateEdit.dateChanged.connect(self.open)
        self.open()

    def openProgram(self):
        self.openProg = True

    def open(self):
        global moyncs
        if str(self.comboBox.currentText()) == 'Январь':
            moyncs = 1
        elif str(self.comboBox.currentText()) == 'Февраль':
            moyncs = 2
        elif str(self.comboBox.currentText()) == 'Март':
            moyncs = 3
        elif str(self.comboBox.currentText()) == 'Апрель':
            moyncs = 4
        elif str(self.comboBox.currentText()) == 'Май':
            moyncs = 5
        elif str(self.comboBox.currentText()) == 'Июнь':
            moyncs = 6
        elif str(self.comboBox.currentText()) == 'Июль':
            moyncs = 7
        elif str(self.comboBox.currentText()) == 'Август':
            moyncs = 8
        elif str(self.comboBox.currentText()) == 'Сентябрь':
            moyncs = 9
        elif str(self.comboBox.currentText()) == 'Октябрь':
            moyncs = 10
        elif str(self.comboBox.currentText()) == 'Ноябрь':
            moyncs = 11
        elif str(self.comboBox.currentText()) == 'Декабрь':
            moyncs = 12
        self.tableWidget.setRowCount(0)
        self.TypeProducts = []
        self.Volume = []
        self.AutoclavaNumber = []
        self.AutoclavaFoo = []
        self.IssuanceManufacture = []
        self.IssuanceManufactureQ = []
        self.IssuanceManufactureA = []
        self.IssuanceTesting = []
        self.IssuanceTestingQ = []
        self.IssuanceTestingA = []
        self.IssuanceProtocol = []
        self.IssuanceProtocolQ = []
        self.IssuanceProtocolA = []
        self.DateSettingCycles = []
        self.DateDeliverySGP = []
        self.Note = []
        self.GT = []
        nameTable = str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy')
        self.cursor.execute('select name from sqlite_master where type = "table"')
        tables = self.cursor.fetchall()
        for i in tables:
            for j in i:
                if j == nameTable:
                    self.cursor.execute(f"SELECT * FROM {nameTable}")
                    j = self.cursor.fetchall()
                    rows = 0
                    for i in j:
                        self.addZapis()
                        if i[3] == 'Плавка':
                            self.Volume[rows].setCurrentIndex(3)
                            self.TypeProducts[rows].setCurrentIndex(7)
                            self.layoutIssuanceManufacture = QHBoxLayout()
                            self.layoutIssuanceManufacture.setContentsMargins(0, 0, 0, 0)
                            self.layoutIssuanceManufacture.setSpacing(0)
                            self.IssuanceManufacture[rows] = QtWidgets.QDateEdit()
                            self.IssuanceManufactureQ[rows] = QtWidgets.QDateEdit()
                            self.IssuanceManufactureA[rows] = QtWidgets.QDateEdit()
                            self.layoutIssuanceManufacture.addWidget(QtWidgets.QLabel('гр.Q'))
                            self.layoutIssuanceManufacture.addWidget(self.IssuanceManufactureQ[rows])
                            self.layoutIssuanceManufacture.addWidget(QtWidgets.QLabel(' хим. а.'))
                            self.layoutIssuanceManufacture.addWidget(self.IssuanceManufactureA[rows])
                            self.IssuanceManufacture[rows].setDisplayFormat('dd.MM')
                            self.IssuanceManufactureQ[rows].setDisplayFormat('dd.MM')
                            self.IssuanceManufactureA[rows].setDisplayFormat('dd.MM')
                            self.IssuanceManufacture[rows].dateChanged.connect(self.selectIssuanceManufacture)
                            self.IssuanceManufactureQ[rows].dateChanged.connect(self.selectIssuanceManufacture)
                            self.IssuanceManufactureA[rows].dateChanged.connect(self.selectIssuanceManufacture)
                            self.CellWidgetlayoutIssuanceManufacture = QWidget()
                            self.CellWidgetlayoutIssuanceManufacture.setLayout(self.layoutIssuanceManufacture)
                            self.tableWidget.setCellWidget(rows, 5, self.CellWidgetlayoutIssuanceManufacture)
                            self.layoutIssuanceTesting = QHBoxLayout()
                            self.layoutIssuanceTesting.setContentsMargins(0, 0, 0, 0)
                            self.layoutIssuanceTesting.setSpacing(0)
                            self.IssuanceTesting[rows] = QtWidgets.QDateEdit()
                            self.IssuanceTestingQ[rows] = QtWidgets.QDateEdit()
                            self.IssuanceTestingA[rows] = QtWidgets.QDateEdit()
                            self.layoutIssuanceTesting.addWidget(QtWidgets.QLabel('гр.Q'))
                            self.layoutIssuanceTesting.addWidget(self.IssuanceTestingQ[rows])
                            self.layoutIssuanceTesting.addWidget(QtWidgets.QLabel(' хим. а.'))
                            self.layoutIssuanceTesting.addWidget(self.IssuanceTestingA[rows])
                            self.IssuanceTesting[rows].setDisplayFormat('dd.MM')
                            self.IssuanceTestingQ[rows].setDisplayFormat('dd.MM')
                            self.IssuanceTestingA[rows].setDisplayFormat('dd.MM')
                            self.CellWidgetlayoutIssuanceTesting = QWidget()
                            self.CellWidgetlayoutIssuanceTesting.setLayout(self.layoutIssuanceTesting)
                            self.tableWidget.setCellWidget(rows, 6, self.CellWidgetlayoutIssuanceTesting)
                            self.layoutIssuanceProtocol = QHBoxLayout()
                            self.layoutIssuanceProtocol.setContentsMargins(0, 0, 0, 0)
                            self.layoutIssuanceProtocol.setSpacing(0)
                            self.IssuanceProtocol[rows] = QtWidgets.QDateEdit()
                            self.IssuanceProtocolQ[rows] = QtWidgets.QDateEdit()
                            self.IssuanceProtocolA[rows] = QtWidgets.QDateEdit()
                            self.layoutIssuanceProtocol.addWidget(QtWidgets.QLabel('гр.Q'))
                            self.layoutIssuanceProtocol.addWidget(self.IssuanceProtocolQ[rows])
                            self.layoutIssuanceProtocol.addWidget(QtWidgets.QLabel(' хим. а.'))
                            self.layoutIssuanceProtocol.addWidget(self.IssuanceProtocolA[rows])
                            self.IssuanceProtocol[rows].setDisplayFormat('dd.MM')
                            self.IssuanceProtocolQ[rows].setDisplayFormat('dd.MM')
                            self.IssuanceProtocolA[rows].setDisplayFormat('dd.MM')
                            self.CellWidgetlayoutIssuanceProtocol = QWidget()
                            self.CellWidgetlayoutIssuanceProtocol.setLayout(self.layoutIssuanceProtocol)
                            self.tableWidget.setCellWidget(rows, 7, self.CellWidgetlayoutIssuanceProtocol)
                            datad = i[6].split(' ')
                            self.IssuanceManufactureQ[rows].setDate(QDate.fromString(datad[0], 'dd.MM'))
                            self.IssuanceManufactureA[rows].setDate(QDate.fromString(datad[1], 'dd.MM'))
                            datad = i[7].split(' ')
                            self.IssuanceTestingQ[rows].setDate(QDate.fromString(datad[0], 'dd.MM'))
                            self.IssuanceTestingA[rows].setDate(QDate.fromString(datad[1], 'dd.MM'))
                            datad = i[8].split(' ')
                            self.IssuanceProtocolQ[rows].setDate(QDate.fromString(datad[0], 'dd.MM'))
                            self.IssuanceProtocolA[rows].setDate(QDate.fromString(datad[1], 'dd.MM'))
                        else:
                            self.IssuanceManufacture[rows].setDate(QDate.fromString(i[6], 'dd.MM'))
                            self.IssuanceTesting[rows].setDate(QDate.fromString(i[7], 'dd.MM'))
                            self.IssuanceProtocol[rows].setDate(QDate.fromString(i[8], 'dd.MM'))
                            self.TypeProducts[rows].setCurrentText(str(i[3]))
                            self.Volume[rows].setCurrentText(str(i[4]))
                        if i[12] == '1':
                            self.GT[rows].setChecked(True)
                        else:
                            self.GT[rows].setChecked(False)
                        self.AutoclavaNumber[rows].setCurrentText(str(i[1]))
                        self.tableWidget.setItem(rows, 1, QTableWidgetItem(str(i[2])))
                        self.AutoclavaFoo[rows].setDate(QDate.fromString(i[5], 'dd.MM'))
                        if i[9] != '':
                            self.DateSettingCycles[rows].setDate(QDate.fromString(i[9], 'dd.MM'))
                        else:
                            self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                        if i[10] != '':
                            self.DateDeliverySGP[rows].setDate(QDate.fromString(i[10], 'dd.MM'))
                        else:
                            self.DateDeliverySGP[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                        if i[11] != '':
                            self.Note[rows].setDate(QDate.fromString(i[11], 'dd.MM'))
                        else:
                            self.Note[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                        rows += 1
        # for i in range(len(self.AutoclavaFoo)):
        #     self.selectH(i)
        self.openProgram()

    def selectIssuanceManufacture(self):
        if self.openProg:
            global moyncs
            combo = self.sender()
            index = self.tableWidget.indexAt(combo.pos())
            fwidget = QApplication.focusWidget()
            if fwidget is not None:
                tableWidget = fwidget.objectName()
            rows = index.row()
            columns = index.column()
            if self.GT[rows].isChecked():
                if self.TypeProducts[rows].currentIndex() == 0:  # 1x25
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 1:  # 1x35
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 2:  # 2x10
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 3:  # 2x15
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 4:  # 2x20x40
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 5:  # 2x50
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 6:  # Затравочное
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 7:  # Плавка
                    pGT = 10
            else:
                pGT = 0
            if self.TypeProducts[rows].currentIndex() == 0:  # 1x25
                qqq = 0
                for i in self.IssuanceManufactureQ:
                    if i == combo:
                        rows = qqq
                        break
                    qqq += 1
                qqq = 0
                for i in self.IssuanceManufactureA:
                    if i == combo:
                        rows = qqq
                        break
                    qqq += 1
                if self.TypeProducts[rows].currentIndex() == 7:
                    dIssuanceTestingQ = www.strptime(self.IssuanceManufactureQ[rows].dateTime().toString('dd.MM'),
                                                     "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                    dIssuanceTestingQ = self.date_by_adding_business_days(dIssuanceTestingQ, 4)  # Выдача на испытания
                    dIssuanceTestingA = www.strptime(self.IssuanceManufactureA[rows].dateTime().toString('dd.MM'),
                                                     "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                    dIssuanceTestingA = self.date_by_adding_business_days(dIssuanceTestingA, 4)  # Выдача на испытания
                    dIssuanceProtocolQ = self.date_by_adding_business_days(dIssuanceTestingQ, 3)  # Выдача протокола
                    dIssuanceProtocolA = self.date_by_adding_business_days(dIssuanceTestingA, 3)  # Выдача протокола
                    dDateSettingCycles = dIssuanceProtocolA + datetime.timedelta(days=5)  # Дата постановки цикла
                    dDateDeliverySGP = dIssuanceProtocolA + datetime.timedelta(days=5)  # Дата сдачи на СГП
                    self.IssuanceTestingQ[rows].setDate(QDate.fromString(dIssuanceTestingQ.strftime("%d.%m"), 'dd.MM'))
                    self.IssuanceTestingA[rows].setDate(QDate.fromString(dIssuanceTestingA.strftime("%d.%m"), 'dd.MM'))
                    self.IssuanceProtocolQ[rows].setDate(QDate.fromString(dIssuanceProtocolQ.strftime("%d.%m"), 'dd.MM'))
                    self.IssuanceProtocolA[rows].setDate(QDate.fromString(dIssuanceProtocolA.strftime("%d.%m"), 'dd.MM'))
                    if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                        self.DateSettingCycles[rows].setDate(
                            QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                    else:
                        self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                    self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
                else:
                    dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                    "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                    dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                    dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                    dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                        days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                    dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(
                        days=5)  # Дата сдачи на СГП (вводится в ручную)
                    self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                    self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                    if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                        self.DateSettingCycles[rows].setDate(
                            QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                    else:
                        self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                    self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 1:  # 1x35
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 2:  # 2x10
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 3:  # 2x15
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 4:  # 2x20x40
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 5:  # 2x50
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 6:  # Затравочное
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 8:  # Л/С
                dIssuanceTesting = www.strptime(self.IssuanceManufacture[rows].dateTime().toString('dd.MM'),
                                                "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceTesting, 4)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 7:  # Плавка
                dIssuanceTestingQ = www.strptime(self.IssuanceManufactureQ[rows].dateTime().toString('dd.MM'),
                                                 "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTestingQ = self.date_by_adding_business_days(dIssuanceTestingQ, 4)  # Выдача на испытания
                dIssuanceTestingA = www.strptime(self.IssuanceManufactureA[rows].dateTime().toString('dd.MM'),
                                                 "%d.%m") + datetime.timedelta(days=1)  # Выдача на испытания
                dIssuanceTestingA = self.date_by_adding_business_days(dIssuanceTestingA, 4)  # Выдача на испытания
                dIssuanceProtocolQ = self.date_by_adding_business_days(dIssuanceTestingQ, 3)  # Выдача протокола
                dIssuanceProtocolA = self.date_by_adding_business_days(dIssuanceTestingA, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocolA + datetime.timedelta(days=5)  # Дата постановки цикла
                dDateDeliverySGP = dIssuanceProtocolA + datetime.timedelta(days=5)  # Дата сдачи на СГП
                self.IssuanceTestingQ[rows].setDate(QDate.fromString(dIssuanceTestingQ.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTestingA[rows].setDate(QDate.fromString(dIssuanceTestingA.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocolQ[rows].setDate(QDate.fromString(dIssuanceProtocolQ.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocolA[rows].setDate(QDate.fromString(dIssuanceProtocolA.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(
                        QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))

    def date_by_adding_days(self, from_date, add_days):
        current_date = from_date + datetime.timedelta(days=add_days)
        current_date = current_date.strftime("%d.%m").split('.')
        d = (calendar.weekday(int(self.dateEdit.dateTime().toString('yyyy')), int(current_date[1]), int(current_date[0])))
        if d == 5:
            current_date = from_date + datetime.timedelta(days=add_days + 2)
        elif d == 6:
            current_date = from_date + datetime.timedelta(days=add_days + 1)
        else:
            current_date = from_date + datetime.timedelta(days=add_days)
        return current_date

    def date_by_adding_business_days(self, from_date, add_days):
        e = 0
        t = from_date.strftime("%d.%m").split('.')
        er = t
        t = (calendar.weekday(int(self.dateEdit.dateTime().toString('yyyy')), int(t[1]), int(t[0])) + 1)
        q = add_days
        www = q
        while q >= 0:
            w = 5 - t
            q -= w
            if q > 0:
                e += 2
                t = 0
        current_date = from_date + datetime.timedelta(days=www + e)
        return current_date

    def selectH(self, rows):
        if self.openProg:
            global moyncs
            if self.GT[rows].isChecked():
                if self.TypeProducts[rows].currentIndex() == 0:  # 1x25
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 1:  # 1x35
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 2:  # 2x10
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 3:  # 2x15
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 4:  # 2x20x40
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 5:  # 2x50
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 6:  # Затравочное
                    pGT = 10
                elif self.TypeProducts[rows].currentIndex() == 7:  # Плавка
                    pGT = 4
            else:
                pGT = 0
            if self.TypeProducts[rows].currentIndex() == 0:  # 1x25
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 12)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 1:  # 1x35
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 13)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 2:  # 2x10
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 8)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 3:  # 2x15
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 14)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 4:  # 2x20x40
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 14)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 5:  # 2x50
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 15)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 6:  # Затравочное
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 8)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 8:  # Л/С
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufacture = self.date_by_adding_days(dAutoclavaFoo, 2)  # Выдача на изготовление
                dIssuanceTesting = self.date_by_adding_business_days(dIssuanceManufacture, 5)  # Выдача на испытания
                dIssuanceProtocol = self.date_by_adding_business_days(dIssuanceTesting, 3)  # Выдача протокола
                dDateSettingCycles = dIssuanceProtocol + datetime.timedelta(
                    days=5 + pGT)  # Дата постановки цикла (вводится в ручную)
                dDateDeliverySGP = dIssuanceProtocol + datetime.timedelta(days=5)  # Дата сдачи на СГП (вводится в ручную)
                self.IssuanceManufacture[rows].setDate(QDate.fromString(dIssuanceManufacture.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTesting[rows].setDate(QDate.fromString(dIssuanceTesting.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocol[rows].setDate(QDate.fromString(dIssuanceProtocol.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))
            elif self.TypeProducts[rows].currentIndex() == 7:  # Плавка
                dAutoclavaFoo = www.strptime(self.AutoclavaFoo[rows].dateTime().toString('dd.MM'), "%d.%m")
                dIssuanceManufactureQ = self.date_by_adding_days(dAutoclavaFoo, 7)  # Выдача на изготовление
                dIssuanceManufactureA = self.date_by_adding_days(dAutoclavaFoo, 10 + pGT)  # Выдача на изготовление
                dIssuanceTestingQ = self.date_by_adding_business_days(dIssuanceManufactureQ, 5)  # Выдача на испытания
                dIssuanceTestingA = self.date_by_adding_business_days(dIssuanceManufactureA, 5)  # Выдача на испытания
                dIssuanceProtocolQ = self.date_by_adding_business_days(dIssuanceTestingQ, 3)  # Выдача протокола
                dIssuanceProtocolA = self.date_by_adding_business_days(dIssuanceTestingA, 3)  # Выдача протокола
                dDateSettingCycles = self.date_by_adding_business_days(dIssuanceProtocolA, 5)  # Дата постановки цикла
                dDateDeliverySGP = self.date_by_adding_business_days(dIssuanceProtocolA, 5)  # Дата сдачи на СГП
                self.IssuanceManufactureQ[rows].setDate(QDate.fromString(dIssuanceManufactureQ.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceManufactureA[rows].setDate(QDate.fromString(dIssuanceManufactureA.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTestingQ[rows].setDate(QDate.fromString(dIssuanceTestingQ.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceTestingA[rows].setDate(QDate.fromString(dIssuanceTestingA.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocolQ[rows].setDate(QDate.fromString(dIssuanceProtocolQ.strftime("%d.%m"), 'dd.MM'))
                self.IssuanceProtocolA[rows].setDate(QDate.fromString(dIssuanceProtocolA.strftime("%d.%m"), 'dd.MM'))
                if int(dDateSettingCycles.strftime("%d.%m").split('.')[1]) <= moyncs:
                    self.DateSettingCycles[rows].setDate(QDate.fromString(dDateSettingCycles.strftime("%d.%m"), 'dd.MM'))
                else:
                    self.DateSettingCycles[rows].setDate(QDate.fromString('01.01', 'dd.MM'))
                self.DateDeliverySGP[rows].setDate(QDate.fromString(dDateDeliverySGP.strftime("%d.%m"), 'dd.MM'))

    def selectAutoclavaFoo(self):
        if self.openProg:
            combo = self.sender()
            index = self.tableWidget.indexAt(combo.pos())
            fwidget = QApplication.focusWidget()
            if fwidget is not None:
                tableWidget = fwidget.objectName()
            rows = index.row()
            columns = index.column()
            # item = combo.currentText()
            self.selectH(rows)

    def selectTypeProducts(self):
        combo = self.sender()
        index = self.tableWidget.indexAt(combo.pos())
        fwidget = QApplication.focusWidget()
        if fwidget is not None:
            tableWidget = fwidget.objectName()
        rows = index.row()
        columns = index.column()
        item = combo.currentText()
        if combo.currentIndex() == 7:
            self.Volume[rows].setCurrentIndex(3)

            self.layoutIssuanceManufacture = QHBoxLayout()
            self.layoutIssuanceManufacture.setContentsMargins(0, 0, 0, 0)
            self.layoutIssuanceManufacture.setSpacing(0)
            self.IssuanceManufacture[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufactureQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufactureA[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufacture[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.IssuanceManufactureQ[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.IssuanceManufactureA[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.layoutIssuanceManufacture.addWidget(QtWidgets.QLabel('гр.Q'))
            self.layoutIssuanceManufacture.addWidget(self.IssuanceManufactureQ[rows])
            self.layoutIssuanceManufacture.addWidget(QtWidgets.QLabel(' хим. а.'))
            self.layoutIssuanceManufacture.addWidget(self.IssuanceManufactureA[rows])
            self.IssuanceManufacture[rows].setDisplayFormat('dd.MM')
            self.IssuanceManufactureQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceManufactureA[rows].setDisplayFormat('dd.MM')
            self.CellWidgetlayoutIssuanceManufacture = QWidget()
            self.CellWidgetlayoutIssuanceManufacture.setLayout(self.layoutIssuanceManufacture)
            self.tableWidget.setCellWidget(rows, 5, self.CellWidgetlayoutIssuanceManufacture)

            self.layoutIssuanceTesting = QHBoxLayout()
            self.layoutIssuanceTesting.setContentsMargins(0, 0, 0, 0)
            self.layoutIssuanceTesting.setSpacing(0)
            self.IssuanceTesting[rows] = QtWidgets.QDateEdit()
            self.IssuanceTestingQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceTestingA[rows] = QtWidgets.QDateEdit()
            self.layoutIssuanceTesting.addWidget(QtWidgets.QLabel('гр.Q'))
            self.layoutIssuanceTesting.addWidget(self.IssuanceTestingQ[rows])
            self.layoutIssuanceTesting.addWidget(QtWidgets.QLabel(' хим. а.'))
            self.layoutIssuanceTesting.addWidget(self.IssuanceTestingA[rows])
            self.IssuanceTesting[rows].setDisplayFormat('dd.MM')
            self.IssuanceTestingQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceTestingA[rows].setDisplayFormat('dd.MM')
            self.CellWidgetlayoutIssuanceTesting = QWidget()
            self.CellWidgetlayoutIssuanceTesting.setLayout(self.layoutIssuanceTesting)
            self.tableWidget.setCellWidget(rows, 6, self.CellWidgetlayoutIssuanceTesting)

            self.layoutIssuanceProtocol = QHBoxLayout()
            self.layoutIssuanceProtocol.setContentsMargins(0, 0, 0, 0)
            self.layoutIssuanceProtocol.setSpacing(0)
            self.IssuanceProtocol[rows] = QtWidgets.QDateEdit()
            self.IssuanceProtocolQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceProtocolA[rows] = QtWidgets.QDateEdit()
            self.layoutIssuanceProtocol.addWidget(QtWidgets.QLabel('гр.Q'))
            self.layoutIssuanceProtocol.addWidget(self.IssuanceProtocolQ[rows])
            self.layoutIssuanceProtocol.addWidget(QtWidgets.QLabel(' хим. а.'))
            self.layoutIssuanceProtocol.addWidget(self.IssuanceProtocolA[rows])
            self.IssuanceProtocol[rows].setDisplayFormat('dd.MM')
            self.IssuanceProtocolQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceProtocolA[rows].setDisplayFormat('dd.MM')
            self.CellWidgetlayoutIssuanceProtocol = QWidget()
            self.CellWidgetlayoutIssuanceProtocol.setLayout(self.layoutIssuanceProtocol)
            self.tableWidget.setCellWidget(rows, 7, self.CellWidgetlayoutIssuanceProtocol)
        else:
            # self.Volume[rows].setCurrentIndex(0)
            self.IssuanceManufactureQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufactureA[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufacture[rows] = QtWidgets.QDateEdit()
            self.IssuanceManufacture[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.IssuanceManufactureQ[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.IssuanceManufactureA[rows].dateChanged.connect(self.selectIssuanceManufacture)
            self.tableWidget.setCellWidget(rows, 5, self.IssuanceManufacture[rows])
            self.IssuanceManufacture[rows].setDisplayFormat('dd.MM')
            self.IssuanceManufactureQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceManufactureA[rows].setDisplayFormat('dd.MM')

            self.IssuanceTestingQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceTestingA[rows] = QtWidgets.QDateEdit()
            self.IssuanceTesting[rows] = QtWidgets.QDateEdit()
            self.tableWidget.setCellWidget(rows, 6, self.IssuanceTesting[rows])
            self.IssuanceTesting[rows].setDisplayFormat('dd.MM')
            self.IssuanceTestingQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceTestingA[rows].setDisplayFormat('dd.MM')

            self.IssuanceProtocolQ[rows] = QtWidgets.QDateEdit()
            self.IssuanceProtocolA[rows] = QtWidgets.QDateEdit()
            self.IssuanceProtocol[rows] = QtWidgets.QDateEdit()
            self.tableWidget.setCellWidget(rows, 7, self.IssuanceProtocol[rows])
            self.IssuanceProtocol[rows].setDisplayFormat('dd.MM')
            self.IssuanceProtocolQ[rows].setDisplayFormat('dd.MM')
            self.IssuanceProtocolA[rows].setDisplayFormat('dd.MM')

    def format(self):
        mesyc = self.comboBox.currentIndex() + 1
        self.save()
        wb = openpyxl.load_workbook('GrafikHablon.xlsx')
        wb.active = 0
        sheet = wb.active
        sheet.title = str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy')
        nameTable = str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy')
        self.cursor.execute(f"SELECT * FROM {nameTable}")
        data = self.cursor.fetchall()
        data = sortMasiv(data)
        h = 13
        n = 12
        r = 1
        Type = data[0][3]
        Volume = data[0][4]
        sheet['A' + str(7)] = 'испытаний на ' + str(
            self.comboBox.currentText()) + ' ' + self.dateEdit.dateTime().toString('yyyy') + ' г.'
        sheet['B' + str(n)] = Type
        sheet['F' + str(n)] = Volume
        for i in data:
            if i[3] != Type or i[4] != Volume:
                if i[3] == 'Плавка':
                    Type = i[3]
                    Volume = i[4]
                    n = h
                    h += 3
                    r = 1
                    sheet.merge_cells('B' + str(n) + ':C' + str(n))
                    sheet.merge_cells('B' + str(n + 1) + ':C' + str(n + 1))
                    sheet.merge_cells('B' + str(n + 2) + ':C' + str(n + 2))
                    sheet.merge_cells('E' + str(n) + ':F' + str(n))
                    sheet.merge_cells('E' + str(n + 1) + ':F' + str(n + 1))
                    sheet.merge_cells('G' + str(n) + ':H' + str(n))
                    sheet.merge_cells('G' + str(n + 1) + ':H' + str(n + 1))
                    sheet.merge_cells('K' + str(n) + ':M' + str(n))
                    sheet.merge_cells('K' + str(n + 1) + ':M' + str(n + 1))
                    font = Font(size=14, bold=True)
                    sheet['B' + str(n)].font = font
                    sheet['B' + str(n + 1)].font = font
                    sheet['B' + str(n)] = i[3]
                    nPlavka = n
                    sheet['B' + str(n + 1)] = i[4]
                    sheet['D' + str(n)] = 'Дата'
                    sheet['D' + str(n + 1)] = 'отключ.'
                    sheet['E' + str(n)] = 'выдача на'
                    sheet['E' + str(n + 1)] = 'изготовление'
                    sheet['E' + str(n + 2)] = 'на гр.Q'
                    sheet['F' + str(n + 2)] = 'на хим.а.'
                    sheet['G' + str(n)] = 'выдача на'
                    sheet['G' + str(n + 1)] = 'испытания'
                    sheet['G' + str(n + 2)] = 'гр.Q'
                    sheet['H' + str(n + 2)] = 'на хим.а.'
                    sheet['K' + str(n)] = 'выдача'
                    sheet['K' + str(n + 1)] = 'протоколов'
                    sheet['K' + str(n + 2)] = 'гр.Q'
                    sheet['M' + str(n + 2)] = 'хим.анализ'
                    sheet['N' + str(n + 2)] = 'дата пост'
                    sheet['O' + str(n + 2)] = 'дата сдачи сгп'
                elif i[3] == 'Л/С':
                    Type = i[3]
                    Volume = i[4]
                    n = h
                    h += 3
                    r = 1
                    sheet.merge_cells('B' + str(n) + ':C' + str(n + 1))
                    sheet.merge_cells('D' + str(n) + ':E' + str(n + 1))
                    sheet.merge_cells('F' + str(n) + ':G' + str(n + 1))
                    sheet.merge_cells('H' + str(n) + ':K' + str(n + 1))
                    sheet.merge_cells('M' + str(n) + ':N' + str(n + 1))
                    sheet.merge_cells('O' + str(n) + ':P' + str(n + 2))
                    font = Font(size=14, bold=True)
                    sheet['B' + str(n)].font = font
                    sheet['B' + str(n)] = i[3]
                    sheet['D' + str(n)] = 'Дата откл.'
                    sheet['F' + str(n)] = 'Выдача на\nизготовление'
                    sheet['H' + str(n)] = 'Выдача на\nиспытания'
                    sheet['M' + str(n)] = 'Выдача\nпротокола'
                    sheet['O' + str(n)] = 'Результат испытаний'
                    sheet['D' + str(n + 2)] = 'план'
                    sheet['E' + str(n + 2)] = 'факт'
                    sheet['F' + str(n + 2)] = 'план'
                    sheet['G' + str(n + 2)] = 'факт'
                    sheet['H' + str(n + 2)] = 'план'
                    sheet['K' + str(n + 2)] = 'факт'
                    sheet['M' + str(n + 2)] = 'план'
                    sheet['N' + str(n + 2)] = 'факт'
                else:
                    Type = i[3]
                    Volume = i[4]
                    n = h
                    h += 1
                    r = 1
                    nPlavka = n
                    font = Font(size=14, bold=True)
                    sheet['B' + str(n)].font = font
                    sheet['F' + str(n)].font = font
                    sheet.merge_cells('B' + str(n) + ':C' + str(n))
                    sheet.merge_cells('F' + str(n) + ':K' + str(n))
                    sheet.merge_cells('M' + str(n) + ':N' + str(n))
                    sheet['B' + str(n)] = i[3]
                    sheet['F' + str(n)] = i[4]
            if i[3] == 'Плавка':
                F = i[6].split(' ')
                sheet['E' + str(h)] = F[0]
                sheet['F' + str(h)] = F[1]
                H = i[7].split(' ')
                sheet['G' + str(h)] = H[0]
                if H[0] != '01.01':
                    if int(H[0].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['G' + str(h)].fill = fill
                sheet['H' + str(h)] = H[1]
                if H[1] != '01.01':
                    if int(H[1].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['H' + str(h)].fill = fill
                K = i[8].split(' ')
                sheet['K' + str(h)] = K[0]
                if K[0] != '01.01':
                    if int(K[0].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['K' + str(h)].fill = fill
                sheet['M' + str(h)] = K[1]
                if K[1] != '01.01':
                    if int(K[1].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['M' + str(h)].fill = fill
                if i[9] != '':
                    sheet['N' + str(h)] = i[9]
                if i[10] != '':
                    sheet['O' + str(h)] = i[10]
                if i[11] != '':
                    sheet['P' + str(h)] = '*график Q-' + str(i[11])
            elif i[3] == 'Л/С':
                sheet['F' + str(h)] = i[6]
                sheet['H' + str(h)] = i[7]
                if i[7] != '01.01':
                    if int(i[7].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['H' + str(h)].fill = fill
                sheet['M' + str(h)] = i[8]
                if i[8] != '01.01':
                    if int(i[8].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['K' + str(h)].fill = fill
                sheet.merge_cells('O' + str(h) + ':P' + str(h))
                # if i[9] != '':
                #     sheet['M' + str(h)] = i[9]
                # if i[10] != '':
                #     sheet['O' + str(h)] = i[10]
                if i[11] != '':
                    sheet['O' + str(h)] = str(i[11])
            else:
                sheet['F' + str(h)] = i[6]
                sheet['H' + str(h)] = i[7]
                if i[7] != '01.01':
                    if int(i[7].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['H' + str(h)].fill = fill
                sheet['K' + str(h)] = i[8]
                if i[8] != '01.01':
                    if int(i[8].split('.')[1].split()[0]) == mesyc:
                        fill = PatternFill(start_color="ff97bb", fill_type="solid")
                        sheet['K' + str(h)].fill = fill
                sheet.merge_cells('M' + str(h) + ':N' + str(h))
                if i[9] != '':
                    sheet['M' + str(h)] = i[9]
                if i[10] != '':
                    sheet['O' + str(h)] = i[10]
                if i[11] != '':
                    sheet['P' + str(h)] = '*график Q-' + str(i[11])
            sheet['A' + str(h)] = r
            sheet['B' + str(h)] = str(i[1]) + '-' + str(i[2])
            sheet['C' + str(h)] = 'zyb/0'
            sheet['D' + str(h)] = i[5]
            h += 1
            r += 1
        u = 13
        nPlavka = n
        for i in range(h - 13):
            sheet['A' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            # if nPlavka != u - 1 and nPlavka != u:
            # print('nPlavka', u)
            sheet['B' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['C' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['D' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['E' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['F' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['G' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['H' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['K' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['M' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            # else:
            #     print('lavka', u)
            #     sheet['B' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['C' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['D' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['E' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['F' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['G' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['H' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['K' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            #     sheet['M' + str(u)].border = Border(
            #         right=Side(border_style='thin', color='FF000000'),
            #         left=Side(border_style='thin', color='FF000000'))
            sheet['N' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['O' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            sheet['P' + str(u)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'),
                                                left=Side(border_style='thin', color='FF000000'))
            u += 1

        sheet.merge_cells('M' + str(u + 1) + ':P' + str(u + 1))
        sheet.merge_cells('M' + str(u + 2) + ':P' + str(u + 2))
        sheet.merge_cells('M' + str(u + 3) + ':P' + str(u + 3))
        sheet.merge_cells('M' + str(u + 6) + ':P' + str(u + 6))
        sheet.merge_cells('M' + str(u + 7) + ':P' + str(u + 7))
        sheet.merge_cells('B' + str(u + 10) + ':C' + str(u + 10))
        sheet.merge_cells('B' + str(u + 11) + ':C' + str(u + 11))
        sheet.merge_cells('D' + str(u + 9) + ':F' + str(u + 9))
        sheet.merge_cells('D' + str(u + 11) + ':F' + str(u + 11))
        sheet.merge_cells('D' + str(u + 13) + ':F' + str(u + 13))
        sheet.merge_cells('D' + str(u + 15) + ':F' + str(u + 15))
        sheet.merge_cells('H' + str(u + 9) + ':M' + str(u + 9))
        sheet.merge_cells('H' + str(u + 11) + ':M' + str(u + 11))
        sheet.merge_cells('H' + str(u + 13) + ':M' + str(u + 13))
        sheet.merge_cells('H' + str(u + 15) + ':M' + str(u + 15))
        sheet.merge_cells('E' + str(u) + ':F' + str(u))
        font = Font(size=12)
        sheet['B' + str(u)].font = font
        sheet['E' + str(u)].font = font
        sheet['B' + str(u)].alignment = Alignment(horizontal="justify")
        sheet['B' + str(u)] = 'ZYB / 0'
        sheet['E' + str(u)] = 'по факту'
        hMesyc = 0
        for i in self.IssuanceTestingQ:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        sheet['G' + str(u)] = hMesyc
        hMesyc = 0
        for i in self.IssuanceTesting:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        for i in self.IssuanceTestingA:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        sheet['H' + str(u)] = hMesyc
        hMesyc = 0
        for i in self.IssuanceProtocol:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        for i in self.IssuanceProtocolQ:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        sheet['K' + str(u)] = hMesyc
        hMesyc = 0
        for i in self.IssuanceProtocolQ:
            if i.dateTime().toString('dd.MM') != '01.01':
                if int(i.dateTime().toString('dd.MM').split('.')[1].split()[0]) == mesyc:
                    hMesyc += 1
        sheet['M' + str(u)] = hMesyc
        font = Font(size=12, bold=True)
        sheet['M' + str(u + 1)].font = font
        sheet['M' + str(u + 2)].font = font
        sheet['M' + str(u + 3)].font = font
        font = Font(size=10, bold=True)
        sheet['M' + str(u + 6)].font = font
        sheet['M' + str(u + 7)].font = font
        font = Font(size=9)
        sheet['B' + str(u + 10)].font = font
        sheet['B' + str(u + 11)].font = font
        font = Font(size=12)
        sheet['D' + str(u + 9)].font = font
        sheet['D' + str(u + 11)].font = font
        sheet['D' + str(u + 13)].font = font
        sheet['D' + str(u + 15)].font = font
        sheet['H' + str(u + 9)].font = font
        sheet['H' + str(u + 11)].font = font
        sheet['H' + str(u + 13)].font = font
        sheet['H' + str(u + 15)].font = font
        sheet['M' + str(u + 1)].alignment = Alignment(horizontal="justify")
        sheet['M' + str(u + 2)].alignment = Alignment(horizontal="justify")
        sheet['M' + str(u + 3)].alignment = Alignment(horizontal="justify")
        sheet['B' + str(u + 10)].alignment = Alignment(horizontal="justify")
        sheet['B' + str(u + 11)].alignment = Alignment(horizontal="justify")
        sheet['D' + str(u + 9)].alignment = Alignment(horizontal="justify")
        sheet['D' + str(u + 11)].alignment = Alignment(horizontal="justify")
        sheet['D' + str(u + 13)].alignment = Alignment(horizontal="justify")
        sheet['D' + str(u + 15)].alignment = Alignment(horizontal="justify")
        sheet['H' + str(u + 9)].alignment = Alignment(horizontal="justify")
        sheet['H' + str(u + 11)].alignment = Alignment(horizontal="justify")
        sheet['H' + str(u + 13)].alignment = Alignment(horizontal="justify")
        sheet['H' + str(u + 15)].alignment = Alignment(horizontal="justify")
        sheet['M' + str(u + 1)] = '* график добротности'
        sheet['M' + str(u + 2)] = 'отправить на электронную почту'
        sheet['M' + str(u + 3)] = 'Главному инженеру и Нач.ТС'
        sheet['M' + str(u + 6)] = '** Затравочные циклы график Q'
        sheet['M' + str(u + 7)] = 'протокол оформлять  в 1 экземпляре для ТС'
        sheet['B' + str(u + 10)] = 'Инженер-технолог ТС'
        sheet['B' + str(u + 11)] = 'Ладынская В.А.'
        sheet['D' + str(u + 9)] = 'Начальник ТС'
        sheet['D' + str(u + 11)] = 'Начальник УСНП'
        sheet['D' + str(u + 13)] = 'Начальник ОТК'
        sheet['D' + str(u + 15)] = 'Начальник ППСО'
        sheet['H' + str(u + 9)] = 'О.В. Михалицына'
        sheet['H' + str(u + 11)] = 'Н.В. Карпинская'
        sheet['H' + str(u + 13)] = 'О.В. Харинская'
        sheet['H' + str(u + 15)] = 'Н.Г. Лапшина'
        wb.save(str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy') + '.xlsx')

    def save(self):
        mmammm = self.comboBox.currentIndex() + 1
        print(mmammm)
        if mmammm == 12:
            mmammm = 13
        nameTable = str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy')
        print(self.dateEdit.dateTime().toString('yyyy'))
        try:
            self.cursor.execute(f"DROP TABLE {nameTable}")
        except:
            self.cursor.execute(
                f'CREATE TABLE {nameTable} (ID, Autoclave_number, Cycle_number, Product_type, Volume, Disconnection_date, '
                f'Issue_for_manufacturing, Issuance_for_testing, Issuance_of_protocol, Cycle_start_date, '
                f'Date_of_submission_SGP, Note, GT)')
        else:
            self.cursor.execute(
                f'CREATE TABLE {nameTable} (ID, Autoclave_number, Cycle_number, Product_type, Volume, Disconnection_date, '
                f'Issue_for_manufacturing, Issuance_for_testing, Issuance_of_protocol, Cycle_start_date, '
                f'Date_of_submission_SGP, Note, GT)')
        nameTable1 = self.comboBox.currentIndex() + 1
        if nameTable1 == 12:
            nameTable1 = 0
        self.comboBox.setCurrentIndex(nameTable1)
        nameTable1 = str(self.comboBox.currentText()) + self.dateEdit.dateTime().toString('yyyy')
        print(nameTable1)
        try:
            self.cursor.execute(f"DROP TABLE {nameTable1}")
        except:
            self.cursor.execute(
                f'CREATE TABLE {nameTable1} (ID, Autoclave_number, Cycle_number, Product_type, Volume, Disconnection_date, '
                f'Issue_for_manufacturing, Issuance_for_testing, Issuance_of_protocol, Cycle_start_date, '
                f'Date_of_submission_SGP, Note, GT)')
        else:
            self.cursor.execute(
                f'CREATE TABLE {nameTable1} (ID, Autoclave_number, Cycle_number, Product_type, Volume, Disconnection_date, '
                f'Issue_for_manufacturing, Issuance_for_testing, Issuance_of_protocol, Cycle_start_date, '
                f'Date_of_submission_SGP, Note, GT)')

        rows = self.tableWidget.rowCount()
        cols = self.tableWidget.columnCount()
        data = []
        # print(self.tableWidget.item(0, 0))
        # for row in range(rows):
        #     tmp = []
        #     for col in range(cols):
        #         try:
        #             tmp.append(self.tableWidget.item(row, col).text())
        #         except:
        #             tmp.append('No data')
        #     data.append(tmp)
        # print(data)
        for i in range(len(self.AutoclavaFoo)):
            try:
                self.Cycle_number = str(self.tableWidget.item(i, 1).text())
            except:
                self.Cycle_number = ''
            if self.DateSettingCycles[i].dateTime().toString('dd.MM') == '01.01':
                self.Cycle_start_date = ''
            else:
                self.Cycle_start_date = self.DateSettingCycles[i].dateTime().toString('dd.MM')
            if self.DateDeliverySGP[i].dateTime().toString('dd.MM') == '01.01':
                self.Date_of_submission_SGP = ''
            else:
                self.Date_of_submission_SGP = self.DateDeliverySGP[i].dateTime().toString('dd.MM')

            if self.Note[i].dateTime().toString('dd.MM') == '01.01':
                self.Date_of_NOTE = ''
            else:
                self.Date_of_NOTE = self.Note[i].dateTime().toString('dd.MM')

            if self.GT[i].isChecked():
                self.bGT = '1'
            else:
                self.bGT = '0'
            if str(self.TypeProducts[i].currentText()) != 'Плавка':
                if int(self.IssuanceTesting[i].dateTime().toString('dd.MM').split('.')[1]) > mmammm:
                    sledm = True
                else:
                    sledm = False
                    '''
                    if 11 > int(self.IssuanceTesting[i].dateTime().toString('dd.MM').split('.')[1]) >= 1:
                        sledm = True
                    '''
                ItemStart = [i + 1, str(self.AutoclavaNumber[i].currentText()), str(self.Cycle_number),
                             str(self.TypeProducts[i].currentText()), str(self.Volume[i].currentText()),
                             self.AutoclavaFoo[i].dateTime().toString('dd.MM'),
                             self.IssuanceManufacture[i].dateTime().toString('dd.MM'),
                             self.IssuanceTesting[i].dateTime().toString('dd.MM'),
                             self.IssuanceProtocol[i].dateTime().toString('dd.MM'),
                             self.Cycle_start_date,
                             self.Date_of_submission_SGP,
                             self.Date_of_NOTE,
                             self.bGT]
            else:
                if int(self.IssuanceTestingQ[i].dateTime().toString('dd.MM').split('.')[1]) > mmammm:
                    sledm = True
                else:
                    sledm = False
                    '''
                    if 11 > int(self.IssuanceTestingQ[i].dateTime().toString('dd.MM').split('.')[1]) >= 1:
                        sledm = True
                    '''
                # print(i + 1, str(self.AutoclavaNumber[i].currentText()), str(self.Cycle_number),
                #              str(self.TypeProducts[i].currentText()), str(self.Volume[i].currentText()),
                #              self.AutoclavaFoo[i].dateTime().toString('dd.MM'),
                #              self.IssuanceManufactureQ[i].dateTime().toString('dd.MM') + ' ' +
                #              self.IssuanceManufactureA[i].dateTime().toString('dd.MM'),
                #              self.IssuanceTestingQ[i].dateTime().toString('dd.MM') + ' ' + self.IssuanceTestingA[
                #                  i].dateTime().toString('dd.MM'),
                #              self.IssuanceProtocolQ[i].dateTime().toString('dd.MM') + ' ' + self.IssuanceProtocolA[
                #                  i].dateTime().toString('dd.MM'),
                #              self.Cycle_start_date,
                #              self.Date_of_submission_SGP,
                #              self.Date_of_NOTE,
                #              self.bGT)
                ItemStart = [i + 1, str(self.AutoclavaNumber[i].currentText()), str(self.Cycle_number),
                             str(self.TypeProducts[i].currentText()), str(self.Volume[i].currentText()),
                             self.AutoclavaFoo[i].dateTime().toString('dd.MM'),
                             self.IssuanceManufactureQ[i].dateTime().toString('dd.MM') + ' ' +
                             self.IssuanceManufactureA[i].dateTime().toString('dd.MM'),
                             self.IssuanceTestingQ[i].dateTime().toString('dd.MM') + ' ' + self.IssuanceTestingA[
                                 i].dateTime().toString('dd.MM'),
                             self.IssuanceProtocolQ[i].dateTime().toString('dd.MM') + ' ' + self.IssuanceProtocolA[
                                 i].dateTime().toString('dd.MM'),
                             self.Cycle_start_date,
                             self.Date_of_submission_SGP,
                             self.Date_of_NOTE,
                             self.bGT]
            if sledm:
                self.cursor.execute(
                    f"INSERT INTO {nameTable1} (ID,Autoclave_number,Cycle_number,Product_type,Volume,Disconnection_date,"
                    f"Issue_for_manufacturing,Issuance_for_testing,Issuance_of_protocol,Cycle_start_date,"
                    f"Date_of_submission_SGP,Note,GT) VALUES {tuple(ItemStart)}")
                self.cursor.execute(
                    f"INSERT INTO {nameTable} (ID,Autoclave_number,Cycle_number,Product_type,Volume,Disconnection_date,"
                    f"Issue_for_manufacturing,Issuance_for_testing,Issuance_of_protocol,Cycle_start_date,"
                    f"Date_of_submission_SGP,Note,GT) VALUES {tuple(ItemStart)}")
            else:
                self.cursor.execute(
                    f"INSERT INTO {nameTable} (ID,Autoclave_number,Cycle_number,Product_type,Volume,Disconnection_date,"
                    f"Issue_for_manufacturing,Issuance_for_testing,Issuance_of_protocol,Cycle_start_date,"
                    f"Date_of_submission_SGP,Note,GT) VALUES {tuple(ItemStart)}")
        self.conn.commit()
        nameTable1 = self.comboBox.currentIndex() - 1
        if nameTable1 == -1:
            nameTable1 = 11
        self.comboBox.setCurrentIndex(nameTable1)

    def addZapis(self):
        self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
        self.TypeProducts.append(QtWidgets.QComboBox())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 2,
                                       self.TypeProducts[self.tableWidget.rowCount() - 1])
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('1x25')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('1x35')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('2x10')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('2x15')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('2x20x40')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('2x50')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('Затравочное')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('Плавка')
        self.TypeProducts[self.tableWidget.rowCount() - 1].addItem('Л/С')
        self.TypeProducts[self.tableWidget.rowCount() - 1].activated.connect(self.selectTypeProducts)
        self.Volume.append(QtWidgets.QComboBox())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 3, self.Volume[self.tableWidget.rowCount() - 1])
        self.Volume[self.tableWidget.rowCount() - 1].addItem('V - 4,0')
        self.Volume[self.tableWidget.rowCount() - 1].addItem('V - 5,0')
        self.Volume[self.tableWidget.rowCount() - 1].addItem('V - 6,0')
        self.Volume[self.tableWidget.rowCount() - 1].addItem('V - 7,0')
        self.AutoclavaNumber.append(QtWidgets.QComboBox())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 0,
                                       self.AutoclavaNumber[self.tableWidget.rowCount() - 1])
        self.cursor.execute("SELECT * FROM Autoclava")
        Autoclava = self.cursor.fetchall()
        for i in Autoclava:
            self.AutoclavaNumber[self.tableWidget.rowCount() - 1].addItem(str(i[1]))
        self.AutoclavaFoo.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 4,
                                       self.AutoclavaFoo[self.tableWidget.rowCount() - 1])
        self.AutoclavaFoo[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')
        self.AutoclavaFoo[self.tableWidget.rowCount() - 1].dateChanged.connect(self.selectAutoclavaFoo)

        self.IssuanceManufacture.append(QtWidgets.QDateEdit())
        self.IssuanceManufactureQ.append(QtWidgets.QDateEdit())
        self.IssuanceManufactureA.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 5,
                                       self.IssuanceManufacture[self.tableWidget.rowCount() - 1])
        self.IssuanceManufacture[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')
        self.IssuanceManufacture[self.tableWidget.rowCount() - 1].dateChanged.connect(self.selectIssuanceManufacture)
        self.IssuanceManufactureQ[self.tableWidget.rowCount() - 1].dateChanged.connect(self.selectIssuanceManufacture)
        self.IssuanceManufactureA[self.tableWidget.rowCount() - 1].dateChanged.connect(self.selectIssuanceManufacture)

        self.IssuanceTesting.append(QtWidgets.QDateEdit())
        self.IssuanceTestingQ.append(QtWidgets.QDateEdit())
        self.IssuanceTestingA.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 6,
                                       self.IssuanceTesting[self.tableWidget.rowCount() - 1])
        self.IssuanceTesting[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')

        self.IssuanceProtocol.append(QtWidgets.QDateEdit())
        self.IssuanceProtocolQ.append(QtWidgets.QDateEdit())
        self.IssuanceProtocolA.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 7,
                                       self.IssuanceProtocol[self.tableWidget.rowCount() - 1])
        self.IssuanceProtocol[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')

        self.DateSettingCycles.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 8,
                                       self.DateSettingCycles[self.tableWidget.rowCount() - 1])
        self.DateSettingCycles[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')

        self.DateDeliverySGP.append(QtWidgets.QDateEdit())
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 9,
                                       self.DateDeliverySGP[self.tableWidget.rowCount() - 1])
        self.DateDeliverySGP[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')

        # Код на два виджета в ячейке
        self.layoutNote = QHBoxLayout()
        self.layoutNote.setContentsMargins(0, 0, 0, 0)
        self.layoutNote.setSpacing(0)
        self.Note.append(QtWidgets.QDateEdit())
        self.layoutNote.addWidget(QtWidgets.QLabel('*график Q-'))
        self.layoutNote.addWidget(self.Note[self.tableWidget.rowCount() - 1])
        self.Note[self.tableWidget.rowCount() - 1].setDisplayFormat('dd.MM')
        self.CellWidgetNote = QWidget()
        self.CellWidgetNote.setLayout(self.layoutNote)
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 10, self.CellWidgetNote)

        self.layoutGT = QHBoxLayout()
        self.layoutGT.setContentsMargins(0, 0, 0, 0)
        self.layoutGT.setAlignment(QtCore.Qt.AlignCenter)
        self.GT.append(QtWidgets.QCheckBox())
        self.layoutGT.addWidget(self.GT[self.tableWidget.rowCount() - 1])
        self.CellWidgetGT = QWidget()
        self.CellWidgetGT.setLayout(self.layoutGT)
        self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, 11,
                                       self.CellWidgetGT)
        self.GT[self.tableWidget.rowCount() - 1].stateChanged.connect(self.selectAutoclavaFoo)

        # for i in self.TypeProducts:
        #     i.activated.connect(self.TypeProductsSelect)

    def delZapis(self):
        try:
            evalSelectedIndexes = self.tableWidget.selectedIndexes()
            for index in evalSelectedIndexes:  # выбранная строка
                rows = index.row()
            self.tableWidget.removeRow(rows)
            self.TypeProducts.pop(rows)
            self.Volume.pop(rows)
            self.AutoclavaNumber.pop(rows)
            self.AutoclavaFoo.pop(rows)
            self.IssuanceManufacture.pop(rows)
            self.IssuanceManufactureQ.pop(rows)
            self.IssuanceManufactureA.pop(rows)
            self.IssuanceTesting.pop(rows)
            self.IssuanceTestingQ.pop(rows)
            self.IssuanceTestingA.pop(rows)
            self.IssuanceProtocol.pop(rows)
            self.IssuanceProtocolQ.pop(rows)
            self.IssuanceProtocolA.pop(rows)
            self.DateSettingCycles.pop(rows)
            self.DateDeliverySGP.pop(rows)
            self.Note.pop(rows)
            self.GT.pop(rows)
        except:
            QMessageBox.about(self, "Ошибка",
                              "Ошибка,  необходимо выбрать строку для удаления!")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
    sys.exit(app.exec_())
